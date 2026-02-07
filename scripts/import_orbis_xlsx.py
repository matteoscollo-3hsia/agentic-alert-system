from __future__ import annotations

import csv
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REQUIRED_HEADERS = {
    "ragione socialecaratteri latini": "name",
    "ragione sociale caratteri latini": "name",
    "numero bvd id": "company_id",
    "paese": "country",
    "nace rev. 2, codici/e primari/o - descrizione": "industry_description",
    "codice nace rev. 2, core code (4 cifre)": "industry_code",
    "totale valore della produzione mil eur ultimo anno disp.": "revenue_mil",
    "indirizzo sito web": "website",
}

LEGAL_SUFFIXES = [
    "S.p.A.",
    "S.p.A",
    "SpA",
    "SPA",
    "S.r.l.",
    "S.r.l",
    "SRL",
    "Srl",
]


def _normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    collapsed = " ".join(text.split())
    return collapsed.strip().lower()


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _parse_revenue(value: Any) -> tuple[str, bool]:
    if value is None or value == "":
        return "", True
    if isinstance(value, (int, float)):
        amount = float(value)
        return str(int(amount * 1_000_000)), False
    text = str(value).strip()
    if not text:
        return "", True
    text = text.replace(",", ".")
    try:
        amount = float(text)
    except ValueError:
        return "", True
    return str(int(amount * 1_000_000)), False


def _normalize_industry_code(value: Any) -> str:
    text = _to_str(value)
    if not text:
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) == 4:
        return digits
    if digits:
        return digits.zfill(4)
    return text


def _normalize_website(value: Any) -> str:
    text = _to_str(value)
    if not text:
        return ""
    if text.startswith("www."):
        return f"https://{text}"
    return text


def _extract_domain(website: str) -> str:
    if not website:
        return ""
    normalized = website.strip()
    if not normalized.startswith(("http://", "https://")):
        normalized = f"https://{normalized}"
    try:
        from urllib.parse import urlparse

        netloc = urlparse(normalized).netloc
    except Exception:
        netloc = normalized
    domain = netloc.lower()
    if domain.startswith("www."):
        domain = domain[4:]
    return domain


def _base_name(name: str) -> str:
    cleaned = name.strip()
    for suffix in LEGAL_SUFFIXES:
        pattern = r"\s+" + re.escape(suffix) + r"$"
        if re.search(pattern, cleaned, flags=re.IGNORECASE):
            cleaned = re.sub(pattern, "", cleaned, flags=re.IGNORECASE).strip()
            break
    return cleaned


def _build_aliases(name: str) -> str:
    aliases: list[str] = []
    seen: set[str] = set()

    def _add(value: str) -> None:
        cleaned = value.strip()
        if not cleaned:
            return
        key = cleaned.lower()
        if key in seen:
            return
        seen.add(key)
        aliases.append(cleaned)

    _add(name)
    base = _base_name(name)
    if base and base.lower() != name.strip().lower():
        _add(base)
    for suffix in LEGAL_SUFFIXES:
        if base:
            _add(f"{base} {suffix}")
    return ";".join(aliases)


def _resolve_input_path() -> Path:
    override = os.getenv("ORBIS_EXPORT_PATH")
    if override:
        return Path(override)
    default_root = Path("orbis_export.xlsx")
    if default_root.exists():
        return default_root
    fallback = Path("data_private") / "orbis_export.xlsx"
    return fallback


def _find_header_row(
    sheet,
    required_keys: set[str],
    max_rows: int = 200,
) -> tuple[int, dict[str, int], int]:
    best_match = 0
    best_header_map: dict[str, int] = {}
    best_row_idx = 1

    for idx, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True),
        start=1,
    ):
        current_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            normalized = _normalize_header(cell)
            if normalized:
                current_map[normalized] = col_idx
        match_count = sum(1 for key in REQUIRED_HEADERS if key in current_map)
        if match_count > best_match:
            best_match = match_count
            best_header_map = current_map
            best_row_idx = idx
        resolved = {
            key: current_map[normalized]
            for normalized, key in REQUIRED_HEADERS.items()
            if normalized in current_map
        }
        if len(set(resolved.keys())) == len(required_keys):
            return idx, current_map, match_count

    return best_row_idx, best_header_map, best_match


def import_orbis_xlsx(
    input_path: Path,
    output_path: Path,
    report_path: Path,
    sheet_name: str | None = None,
) -> dict[str, int]:
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)

    required_keys = set(REQUIRED_HEADERS.values())

    if sheet_name:
        sheet = workbook[sheet_name]
        header_row_idx, header_map, _ = _find_header_row(
            sheet, required_keys
        )
    else:
        best_sheet = workbook.worksheets[0]
        header_row_idx, header_map, best_match = _find_header_row(
            best_sheet, required_keys
        )
        for candidate in workbook.worksheets[1:]:
            row_idx, candidate_map, candidate_match = _find_header_row(
                candidate, required_keys
            )
            if candidate_match > best_match:
                best_match = candidate_match
                best_sheet = candidate
                header_row_idx = row_idx
                header_map = candidate_map
        sheet = best_sheet

    required = {
        key: header_map[normalized]
        for normalized, key in REQUIRED_HEADERS.items()
        if normalized in header_map
    }

    missing = [key for key in required_keys if key not in required]
    if missing:
        missing_labels = ", ".join(sorted(missing))
        raise ValueError(
            f"Missing required headers: {missing_labels} (sheet: {sheet.title})"
        )

    rows_in = 0
    rows_written = 0
    missing_website = 0
    missing_revenue = 0
    dropped: list[dict[str, str]] = []

    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", newline="", encoding="utf-8") as out_handle:
        writer = csv.DictWriter(
            out_handle,
            fieldnames=[
                "company_id",
                "name",
                "aliases",
                "revenue_eur",
                "industry_code",
                "industry_description",
                "website",
                "website_domain",
                "country",
                "contact_owner",
                "status",
            ],
        )
        writer.writeheader()

        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if row is None:
                continue
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            rows_in += 1
            company_id = _to_str(row[required["company_id"]]).strip()
            name = _to_str(row[required["name"]]).strip()
            if not company_id or not name:
                reason = []
                if not company_id:
                    reason.append("missing_company_id")
                if not name:
                    reason.append("missing_name")
                dropped.append(
                    {
                        "company_id": company_id,
                        "name": name,
                        "reason": "+".join(reason),
                    }
                )
                continue

            country = _to_str(row[required["country"]]).strip()
            industry_description = _to_str(
                row[required["industry_description"]]
            ).strip()
            industry_code = _normalize_industry_code(
                row[required["industry_code"]]
            )

            revenue_eur, revenue_missing = _parse_revenue(
                row[required["revenue_mil"]]
            )
            if revenue_missing:
                missing_revenue += 1

            website = _normalize_website(row[required["website"]])
            if not website:
                missing_website += 1
            website_domain = _extract_domain(website)

            aliases = _build_aliases(name)

            writer.writerow(
                {
                    "company_id": company_id,
                    "name": name,
                    "aliases": aliases,
                    "revenue_eur": revenue_eur,
                    "industry_code": industry_code,
                    "industry_description": industry_description,
                    "website": website,
                    "website_domain": website_domain,
                    "country": country,
                    "contact_owner": "N/A",
                    "status": "active",
                }
            )
            rows_written += 1

    with report_path.open("w", newline="", encoding="utf-8") as report_handle:
        writer = csv.DictWriter(
            report_handle, fieldnames=["company_id", "name", "reason"]
        )
        writer.writeheader()
        for row in dropped:
            writer.writerow(row)

    return {
        "rows_in": rows_in,
        "rows_written": rows_written,
        "missing_website": missing_website,
        "missing_revenue": missing_revenue,
    }


def main() -> None:
    input_path = _resolve_input_path()
    sheet_name = os.getenv("ORBIS_SHEET_NAME")
    output_path = Path("data_private") / "companies.csv"
    report_path = Path("data_private") / "import_orbis_report.csv"

    summary = import_orbis_xlsx(
        input_path=input_path,
        output_path=output_path,
        report_path=report_path,
        sheet_name=sheet_name,
    )

    print(
        "rows_in={rows_in} rows_written={rows_written} "
        "missing_website={missing_website} missing_revenue={missing_revenue}".format(
            **summary
        )
    )


if __name__ == "__main__":
    main()
